"""Strict XLSX parsing for the administration operations centre.

The public scoreline and school-notice datasets have different shapes, but
their import safety requirements are identical: a fixed spreadsheet contract,
row-level diagnostics, duplicate detection before writes, and a stable file
digest for version history.  Keeping this code independent of FastAPI makes
the preview and commit routes use exactly the same validation path.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook


MAX_IMPORT_BYTES = 8 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000
MAX_MAJOR_CATALOG_IMPORT_ROWS = 60_000
SUPPORTED_OPERATION_DATASETS = {"scorelines", "announcements", "major-catalog"}


class OperationsImportError(ValueError):
    """Raised for a file-level spreadsheet error."""


@dataclass(frozen=True)
class ParsedImport:
    dataset: str
    source_sha256: str
    records: list[dict[str, Any]]
    errors: list[dict[str, Any]]
    warnings: list[str]

    @property
    def total_rows(self) -> int:
        return len(self.records) + len(self.errors)

    @property
    def is_valid(self) -> bool:
        return not self.errors and bool(self.records)


def _normalized_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold().strip()
    return re.sub(r"[\s_\-()（）/\\]+", "", text)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _header_map(headers: list[Any], aliases: dict[str, set[str]]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    normalized_aliases = {
        field: {_normalized_header(alias) for alias in field_aliases}
        for field, field_aliases in aliases.items()
    }
    for index, header in enumerate(headers):
        normalized = _normalized_header(header)
        if not normalized:
            continue
        for field, field_aliases in normalized_aliases.items():
            if field not in resolved and normalized in field_aliases:
                resolved[field] = index
                break
    return resolved


def _load_rows(content: bytes, *, max_rows: int = MAX_IMPORT_ROWS) -> list[tuple[int, list[Any]]]:
    if not content:
        raise OperationsImportError("上传文件为空")
    if len(content) > MAX_IMPORT_BYTES:
        raise OperationsImportError("XLSX 文件超过 8MB 限制")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise OperationsImportError("文件不是可读取的 XLSX 工作簿") from exc

    try:
        worksheet = workbook.active
        rows = [(index, list(values)) for index, values in enumerate(worksheet.iter_rows(values_only=True), start=1)]
    finally:
        workbook.close()
    if not rows:
        raise OperationsImportError("工作表为空")
    if len(rows) - 1 > max_rows:
        raise OperationsImportError(f"单次最多导入 {max_rows} 行")
    return rows


def _value(row: list[Any], mapping: dict[str, int], field: str) -> str:
    index = mapping.get(field)
    return _cell_text(row[index]) if index is not None and index < len(row) else ""


def _year(value: str) -> str:
    normalized = value.strip()
    if re.fullmatch(r"20\d{2}", normalized):
        return normalized
    return ""


def _number(value: str) -> float | None:
    normalized = value.replace(",", "").strip()
    if not normalized:
        return None
    if re.fullmatch(r"\d+(?:\.\d+)?", normalized):
        return float(normalized)
    return None


SCORELINE_HEADERS = {
    "score_year": {"年份", "年度", "year", "score_year", "score year"},
    "region": {"地区", "省份", "区域", "region", "province"},
    "school_name": {"院校", "学校", "招生单位", "school", "school_name"},
    "unit_name": {"院系", "学院", "培养单位", "单位", "unit", "department", "unit_name"},
    "score_raw": {"分数线", "原始分数", "分数", "score", "score_raw"},
    "score_kind": {"分数类型", "score_kind", "状态"},
    "source_url": {"来源链接", "链接", "网址", "source_url", "url"},
    "source_note": {"说明", "来源说明", "备注", "source_note", "note"},
}

ANNOUNCEMENT_HEADERS = {
    "notice_year": {"年份", "年度", "year", "notice_year"},
    "region": {"地区", "省份", "区域", "region", "province"},
    "school_name": {"院校", "学校", "招生单位", "school", "school_name"},
    "unit_name": {"院系", "学院", "培养单位", "单位", "unit", "department", "unit_name"},
    "notice_type": {"公告类型", "类型", "notice_type", "type"},
    "title": {"标题", "公告标题", "title"},
    "summary": {"摘要", "简介", "summary"},
    "notice_date": {"发布日期", "公告日期", "日期", "notice_date", "date"},
    "source_url": {"来源链接", "链接", "网址", "source_url", "url"},
    "content_text": {"正文", "公告正文", "内容", "content", "content_text"},
}

MAJOR_CATALOG_HEADERS = {
    "catalog_year": {"年份", "目录年份", "年度", "catalog_year", "year"},
    "region": {"地区", "省份", "区域", "region", "province"},
    "school_name": {"院校", "学校", "招生单位", "school", "school_name"},
    "department_name": {"院系", "学院", "培养单位", "院系所", "department", "department_name"},
    "program_name": {"专业", "专业名称", "program", "program_name"},
    "program_code": {"专业代码", "代码", "program_code", "code"},
    "direction_name": {"研究方向", "方向", "direction", "direction_name"},
    "tutor": {"导师", "指导教师", "tutor"},
    "exam_code": {"考试类别", "考试代码", "考试类型", "exam_code", "exam"},
    "degree": {"学位", "学位类型", "degree"},
    "study_mode": {"学习方式", "培养方式", "study_mode", "mode"},
}

REQUIRED_SCORELINE_FIELDS = ("score_year", "region", "school_name")
REQUIRED_ANNOUNCEMENT_FIELDS = ("notice_year", "region", "school_name", "notice_type", "title")
REQUIRED_MAJOR_CATALOG_FIELDS = ("catalog_year", "region", "school_name", "program_name", "exam_code")


def _require_headers(mapping: dict[str, int], required: tuple[str, ...]) -> None:
    missing = [field for field in required if field not in mapping]
    if missing:
        raise OperationsImportError("缺少必填列：" + "、".join(missing))


def _locate_header(
    rows: list[tuple[int, list[Any]]],
    aliases: dict[str, set[str]],
    required: tuple[str, ...],
) -> tuple[int, dict[str, int], list[tuple[int, list[Any]]]]:
    """Accept a plain first-row header and the labelled templates used by ops."""
    fallback_mapping: dict[str, int] = {}
    for index, (row_number, headers) in enumerate(rows[:10]):
        mapping = _header_map(headers, aliases)
        if len(mapping) > len(fallback_mapping):
            fallback_mapping = mapping
        if all(field in mapping for field in required):
            return row_number, mapping, rows[index + 1:]
    _require_headers(fallback_mapping, required)
    raise OperationsImportError("未找到字段表头")


def _score_kind(value: str, raw: str, numeric: float | None) -> str:
    normalized = value.casefold().strip()
    aliases = {
        "score": "score",
        "分数": "score",
        "missing": "missing",
        "缺失": "missing",
        "unavailable": "unavailable",
        "暂无": "unavailable",
        "official": "official",
        "官方": "official",
        "multiple": "multiple",
        "多个": "multiple",
        "note": "note",
        "备注": "note",
    }
    if normalized in aliases:
        return aliases[normalized]
    if numeric is not None:
        return "score"
    if not raw:
        return "missing"
    if any(token in raw for token in ("暂无", "未公布", "无数据", "待公布")):
        return "unavailable"
    return "note"


def parse_scoreline_xlsx(content: bytes) -> ParsedImport:
    rows = _load_rows(content)
    header_row_number, mapping, data_rows = _locate_header(rows, SCORELINE_HEADERS, REQUIRED_SCORELINE_FIELDS)
    if "score_raw" not in mapping:
        raise OperationsImportError("历年分数线需要“分数线”列")

    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    duplicate_keys: set[tuple[str, ...]] = set()
    for row_number, row in data_rows:
        if not any(_cell_text(value) for value in row):
            continue
        year = _year(_value(row, mapping, "score_year"))
        region = _value(row, mapping, "region")
        school_name = _value(row, mapping, "school_name")
        unit_name = _value(row, mapping, "unit_name")
        raw = _value(row, mapping, "score_raw")
        numeric = _number(raw)
        row_errors: list[str] = []
        if not year:
            row_errors.append("年份必须为 20xx")
        if not region:
            row_errors.append("地区不能为空")
        if not school_name:
            row_errors.append("院校不能为空")
        if not raw and numeric is None:
            row_errors.append("分数不能为空")
        key = (year, region, school_name, unit_name, raw, _value(row, mapping, "source_url"))
        if not row_errors and key in duplicate_keys:
            row_errors.append("与本文件中的另一行重复")
        if row_errors:
            errors.append({"row": row_number, "errors": row_errors})
            continue
        duplicate_keys.add(key)
        records.append({
            "score_year": year,
            "region": region,
            "school_name": school_name,
            "unit_name": unit_name,
            "score_raw": raw or str(numeric),
            "score_value": numeric,
            "score_kind": _score_kind(_value(row, mapping, "score_kind"), raw, numeric),
            "source_url": _value(row, mapping, "source_url") or None,
            "source_note": _value(row, mapping, "source_note") or None,
            "source_row": row_number,
        })

    if not records and not errors:
        raise OperationsImportError(f"第 {header_row_number} 行表头后没有可导入数据")
    return ParsedImport("scorelines", sha256(content).hexdigest(), records, errors, [])


def _notice_type(value: str) -> str:
    normalized = value.casefold().strip()
    aliases = {
        "brochure": "brochure",
        "招生简章": "brochure",
        "简章": "brochure",
        "scoreline_retest": "scoreline_retest",
        "复试分数线": "scoreline_retest",
        "分数线": "scoreline_retest",
    }
    return aliases.get(normalized, "")


def parse_school_announcement_xlsx(content: bytes) -> ParsedImport:
    rows = _load_rows(content)
    header_row_number, mapping, data_rows = _locate_header(rows, ANNOUNCEMENT_HEADERS, REQUIRED_ANNOUNCEMENT_FIELDS)

    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    duplicate_keys: set[tuple[str, ...]] = set()
    for row_number, row in data_rows:
        if not any(_cell_text(value) for value in row):
            continue
        year = _year(_value(row, mapping, "notice_year"))
        region = _value(row, mapping, "region")
        school_name = _value(row, mapping, "school_name")
        unit_name = _value(row, mapping, "unit_name")
        notice_type = _notice_type(_value(row, mapping, "notice_type"))
        title = _value(row, mapping, "title")
        content_text = _value(row, mapping, "content_text")
        summary = _value(row, mapping, "summary") or content_text[:300]
        row_errors: list[str] = []
        if not year:
            row_errors.append("年份必须为 20xx")
        if not region:
            row_errors.append("地区不能为空")
        if not school_name:
            row_errors.append("院校不能为空")
        if not notice_type:
            row_errors.append("公告类型仅支持 招生简章 或 复试分数线")
        if not title:
            row_errors.append("标题不能为空")
        key = (year, region, school_name, unit_name, notice_type, title, _value(row, mapping, "source_url"))
        if not row_errors and key in duplicate_keys:
            row_errors.append("与本文件中的另一行重复")
        if row_errors:
            errors.append({"row": row_number, "errors": row_errors})
            continue
        duplicate_keys.add(key)
        records.append({
            "notice_year": year,
            "region": region,
            "school_name": school_name,
            "unit_name": unit_name,
            "notice_type": notice_type,
            "title": title,
            "summary": summary,
            "notice_date": _value(row, mapping, "notice_date") or None,
            "source_url": _value(row, mapping, "source_url") or None,
            "content_text": content_text,
            "source_row": row_number,
        })

    if not records and not errors:
        raise OperationsImportError(f"第 {header_row_number} 行表头后没有可导入数据")
    return ParsedImport("announcements", sha256(content).hexdigest(), records, errors, [])


def parse_major_catalog_xlsx(content: bytes) -> ParsedImport:
    rows = _load_rows(content, max_rows=MAX_MAJOR_CATALOG_IMPORT_ROWS)
    header_row_number, mapping, data_rows = _locate_header(rows, MAJOR_CATALOG_HEADERS, REQUIRED_MAJOR_CATALOG_FIELDS)
    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    duplicate_keys: set[tuple[str, ...]] = set()
    years: set[str] = set()
    for row_number, row in data_rows:
        if not any(_cell_text(value) for value in row):
            continue
        catalog_year = _year(_value(row, mapping, "catalog_year"))
        region = _value(row, mapping, "region")
        school_name = _value(row, mapping, "school_name")
        department_name = _value(row, mapping, "department_name") or "未区分院系所"
        program_name = _value(row, mapping, "program_name")
        program_code = _value(row, mapping, "program_code")
        direction_name = _value(row, mapping, "direction_name") or "不区分研究方向"
        tutor = _value(row, mapping, "tutor")
        exam_code = _value(row, mapping, "exam_code").upper()
        degree = _value(row, mapping, "degree")
        study_mode = _value(row, mapping, "study_mode")
        row_errors: list[str] = []
        if not catalog_year:
            row_errors.append("年份必须为 20xx")
        if not region:
            row_errors.append("地区不能为空")
        if not school_name:
            row_errors.append("院校不能为空")
        if not program_name:
            row_errors.append("专业不能为空")
        if exam_code not in {"Z001", "Z002"}:
            row_errors.append("考试类别仅支持 Z001 或 Z002")
        key = (
            catalog_year,
            region,
            school_name,
            department_name,
            program_name,
            program_code,
            direction_name,
            tutor,
            exam_code,
            degree,
            study_mode,
        )
        if not row_errors and key in duplicate_keys:
            row_errors.append("与本文件中的另一行重复")
        if row_errors:
            errors.append({"row": row_number, "errors": row_errors})
            continue
        duplicate_keys.add(key)
        years.add(catalog_year)
        records.append({
            "catalog_year": catalog_year,
            "region": region,
            "school_name": school_name,
            "department_name": department_name,
            "program_name": program_name,
            "program_code": program_code,
            "direction_name": direction_name,
            "tutor": tutor,
            "exam_code": exam_code,
            "degree": degree,
            "study_mode": study_mode,
            "source_row": row_number,
        })

    if not records and not errors:
        raise OperationsImportError(f"第 {header_row_number} 行表头后没有可导入数据")
    if len(years) > 1:
        errors.append({"row": header_row_number, "errors": ["一个专业目录文件只能包含一个年份"]})
    return ParsedImport("major-catalog", sha256(content).hexdigest(), records, errors, [])


def parse_operations_xlsx(dataset: str, content: bytes) -> ParsedImport:
    """Parse one supported operations workbook through the shared contract."""
    normalized_dataset = (dataset or "").strip().lower()
    if normalized_dataset == "scorelines":
        return parse_scoreline_xlsx(content)
    if normalized_dataset == "announcements":
        return parse_school_announcement_xlsx(content)
    if normalized_dataset == "major-catalog":
        return parse_major_catalog_xlsx(content)
    raise OperationsImportError("不支持的数据类型")


def import_run_statistics(parsed: ParsedImport) -> dict[str, int | str]:
    """Return only stable, audit-safe summary values for an import run."""
    summary: dict[str, int | str] = {
        "dataset": parsed.dataset,
        "total_rows": parsed.total_rows,
        "valid_rows": len(parsed.records),
        "invalid_rows": len(parsed.errors),
    }
    if parsed.dataset == "major-catalog" and parsed.records:
        summary["catalog_year"] = str(parsed.records[0].get("catalog_year") or "")
    return summary


def import_preview_items(parsed: ParsedImport, limit: int = 100) -> list[dict[str, Any]]:
    """Build a compact preview without exposing the entire uploaded workbook."""
    items: list[dict[str, Any]] = []
    for record in parsed.records[:limit]:
        row = dict(record)
        row["valid"] = True
        items.append(row)
    remaining = max(0, limit - len(items))
    for error in parsed.errors[:remaining]:
        items.append({
            "source_row": error.get("row"),
            "valid": False,
            "errors": error.get("errors") or [],
        })
    return items


def build_import_records(parsed: ParsedImport, import_run_id: str) -> list[dict[str, Any]]:
    """Convert validated parser rows into the exact database insert shape."""
    if parsed.dataset == "scorelines":
        return [
            {
                "import_run_id": import_run_id,
                "score_year": record["score_year"],
                "region": record["region"],
                "school_name": record["school_name"],
                "unit_name": record["unit_name"],
                "score_raw": record["score_raw"],
                "score_value": record["score_value"],
                "score_kind": record["score_kind"],
                "source_url": record["source_url"],
                "source_note": record["source_note"],
                "is_published": False,
            }
            for record in parsed.records
        ]
    if parsed.dataset == "announcements":
        return [
            {
                "import_run_id": import_run_id,
                "notice_year": record["notice_year"],
                "region": record["region"],
                "school_name": record["school_name"],
                "unit_name": record["unit_name"],
                "notice_type": record["notice_type"],
                "title": record["title"],
                "summary": record["summary"],
                "notice_date": record["notice_date"],
                "source_url": record["source_url"],
                "content_text": record["content_text"],
                "is_published": False,
                "status": "draft",
            }
            for record in parsed.records
        ]
    if parsed.dataset == "major-catalog":
        return [
            {
                "import_run_id": import_run_id,
                "catalog_year": record["catalog_year"],
                "region": record["region"],
                "school_name": record["school_name"],
                "department_name": record["department_name"],
                "program_name": record["program_name"],
                "program_code": record["program_code"],
                "direction_name": record["direction_name"],
                "tutor": record["tutor"],
                "exam_code": record["exam_code"],
                "degree": record["degree"],
                "study_mode": record["study_mode"],
                "source_row": record["source_row"],
            }
            for record in parsed.records
        ]
    raise OperationsImportError("不支持的数据类型")
